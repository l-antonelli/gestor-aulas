"""Tests para plan_grilla_export_service: export a Excel del cronograma
del plan (metadata + cronograma visual + detalle plano)."""

from datetime import time
from io import BytesIO

from openpyxl import load_workbook

from src.services.plan_grilla_export_service import (
    build_export_filename,
    export_grilla_a_xlsx,
)
from src.services.schedule_service import ScheduleBlock


def _mk_block(
    entry_id: str,
    materia_codigo: str = "MAT1",
    materia_nombre: str = "Matemática 1",
    hi: int = 8,
    hf: int = 10,
    comision_numero: int | None = 1,
    aula_label: str | None = "Pellegrini · AULA 10",
    virtual: bool = False,
    tipo_clase: str | None = "teorica",
    carreras_label: str | None = None,
) -> ScheduleBlock:
    return ScheduleBlock(
        entry_id=entry_id,
        materia_codigo=materia_codigo,
        materia_nombre=materia_nombre,
        hora_inicio=time(hi, 0),
        hora_fin=time(hf, 0),
        comision_id=None,
        comision_numero=comision_numero,
        comision_nombre=None,
        aula_label=aula_label,
        virtual=virtual,
        tipo_clase=tipo_clase,
        carreras_label=carreras_label,
    )


class TestExportGrillaAXlsx:

    def test_genera_xlsx_valido_con_tres_hojas(self):
        grid = {"Lunes": [_mk_block("h1")]}
        raw = export_grilla_a_xlsx(
            grid_data=grid,
            plan_nombre="Plan Test",
            ciclo_label="2026-1C",
            filtros={"Carrera": "A", "Año": "3º", "Cuatrimestre": "1C"},
        )
        wb = load_workbook(BytesIO(raw))
        assert wb.sheetnames == ["Metadata", "Cronograma", "Detalle"]

    def test_metadata_incluye_filtros(self):
        grid = {"Lunes": [_mk_block("h1")]}
        raw = export_grilla_a_xlsx(
            grid_data=grid,
            plan_nombre="Plan Test",
            ciclo_label="2026-1C",
            filtros={
                "Carrera": "A · Electrónica",
                "Alcance": "Sólo comunes",
            },
        )
        wb = load_workbook(BytesIO(raw))
        ws = wb["Metadata"]
        # Buscar la celda con "Plan"; la de al lado debe ser el nombre.
        contenidos = {
            ws.cell(row=r, column=1).value: ws.cell(row=r, column=2).value
            for r in range(1, ws.max_row + 1)
        }
        assert contenidos.get("Plan") == "Plan Test"
        assert contenidos.get("Ciclo") == "2026-1C"
        assert contenidos.get("Carrera") == "A · Electrónica"
        assert contenidos.get("Alcance") == "Sólo comunes"

    def test_detalle_tiene_fila_por_bloque(self):
        grid = {
            "Lunes": [
                _mk_block("h1", materia_codigo="MAT1"),
                _mk_block("h2", materia_codigo="MAT2"),
            ],
            "Martes": [_mk_block("h3", materia_codigo="MAT3")],
        }
        raw = export_grilla_a_xlsx(
            grid_data=grid,
            plan_nombre="P",
            ciclo_label="C",
            filtros={},
        )
        wb = load_workbook(BytesIO(raw))
        ws = wb["Detalle"]
        # Header + 3 filas de datos.
        assert ws.max_row == 4
        codigos = {ws.cell(row=r, column=1).value for r in range(2, 5)}
        assert codigos == {"MAT1", "MAT2", "MAT3"}

    def test_detalle_muestra_modalidad_y_aula(self):
        grid = {
            "Lunes": [
                _mk_block(
                    "h1", virtual=True, aula_label=None,
                    tipo_clase="teorica",
                ),
                _mk_block(
                    "h2", virtual=False,
                    aula_label="Pellegrini · AULA 5",
                    tipo_clase="laboratorio",
                ),
                _mk_block(
                    "h3", virtual=False, aula_label=None,
                    tipo_clase=None,
                ),
            ],
        }
        raw = export_grilla_a_xlsx(
            grid_data=grid, plan_nombre="P", ciclo_label="C",
            filtros={},
        )
        wb = load_workbook(BytesIO(raw))
        ws = wb["Detalle"]
        # Columnas: 1=cod, 2=nombre, 3=com, 4=alcance, 5=dia,
        # 6=hi, 7=hf, 8=aula, 9=modalidad
        aulas = [ws.cell(row=r, column=8).value for r in range(2, 5)]
        modalidades = [
            ws.cell(row=r, column=9).value for r in range(2, 5)
        ]
        assert "Virtual" in aulas
        assert "Pellegrini · AULA 5" in aulas
        assert "Sin aula" in aulas
        assert "Virtual" in modalidades
        assert "Laboratorio" in modalidades

    def test_detalle_muestra_alcance_desde_carreras_label(self):
        grid = {
            "Lunes": [
                _mk_block(
                    "h1", materia_codigo="F14",
                    carreras_label="Común (A, E, M)",
                ),
                _mk_block(
                    "h2", materia_codigo="A20",
                    carreras_label="A",
                ),
            ],
        }
        raw = export_grilla_a_xlsx(
            grid_data=grid, plan_nombre="P", ciclo_label="C",
            filtros={},
        )
        wb = load_workbook(BytesIO(raw))
        ws = wb["Detalle"]
        alcances = {
            ws.cell(row=r, column=1).value: ws.cell(row=r, column=4).value
            for r in range(2, ws.max_row + 1)
        }
        assert alcances["F14"] == "Común (A, E, M)"
        assert alcances["A20"] == "A"

    def test_grid_vacio_produce_hoja_detalle_solo_con_header(self):
        raw = export_grilla_a_xlsx(
            grid_data={}, plan_nombre="P", ciclo_label="C",
            filtros={},
        )
        wb = load_workbook(BytesIO(raw))
        assert wb["Detalle"].max_row == 1  # sólo header


class TestCronogramaSheetEstetica:
    """Verifica los cambios estéticos de la hoja Cronograma:
    - bloques merged verticalmente cuando duran múltiples slots,
    - color por materia,
    - sub-columnas por día cuando hay clases paralelas.
    """

    def test_bloque_de_dos_horas_merged_verticalmente(self):
        # Bloque 10:00 → 12:00 = 8 slots de 15 min. Debería quedar
        # merged en una única celda que abarque 8 filas.
        grid = {
            "Lunes": [
                _mk_block("h1", materia_codigo="MAT1", hi=10, hf=12),
            ],
        }
        raw = export_grilla_a_xlsx(
            grid_data=grid, plan_nombre="P", ciclo_label="C",
            filtros={},
        )
        wb = load_workbook(BytesIO(raw))
        ws = wb["Cronograma"]

        # Encontrar un merge que abarque 8 filas en la columna B.
        merges_en_col_b = [
            m for m in ws.merged_cells.ranges
            if m.min_col == 2 and m.max_col == 2
        ]
        # Al menos un merge de 8 filas verticales.
        alturas = [
            m.max_row - m.min_row + 1 for m in merges_en_col_b
        ]
        assert 8 in alturas, (
            f"esperaba un merge de 8 filas en col B; got {alturas}"
        )

    def test_dia_con_clases_paralelas_usa_subcolumnas(self):
        # Dos bloques simultáneos en Lunes → dos lanes = 2 columnas
        # para Lunes.
        grid = {
            "Lunes": [
                _mk_block("h1", materia_codigo="MAT1", hi=10, hf=12),
                _mk_block("h2", materia_codigo="MAT2", hi=10, hf=12),
            ],
        }
        raw = export_grilla_a_xlsx(
            grid_data=grid, plan_nombre="P", ciclo_label="C",
            filtros={},
        )
        wb = load_workbook(BytesIO(raw))
        ws = wb["Cronograma"]

        # Header (fila 1): la celda "Lunes" debe estar mergeada
        # sobre 2 columnas (B y C).
        merges_lunes_header = [
            m for m in ws.merged_cells.ranges
            if m.min_row == 1 and m.max_row == 1 and m.min_col == 2
        ]
        assert merges_lunes_header, (
            "Header de Lunes debería estar mergeado sobre >1 col"
        )
        m = merges_lunes_header[0]
        assert m.max_col - m.min_col + 1 == 2, (
            f"esperaba 2 columnas bajo Lunes; got "
            f"{m.max_col - m.min_col + 1}"
        )
        # El valor del header viene en la primera celda del merge.
        assert ws.cell(row=1, column=m.min_col).value == "Lunes"

    def test_dia_sin_paralelismo_usa_una_columna(self):
        # Tres bloques SECUENCIALES (no simultáneos) en Lunes →
        # entran en un solo lane.
        grid = {
            "Lunes": [
                _mk_block("h1", materia_codigo="MAT1", hi=8, hf=10),
                _mk_block("h2", materia_codigo="MAT2", hi=10, hf=12),
                _mk_block("h3", materia_codigo="MAT3", hi=13, hf=15),
            ],
        }
        raw = export_grilla_a_xlsx(
            grid_data=grid, plan_nombre="P", ciclo_label="C",
            filtros={},
        )
        wb = load_workbook(BytesIO(raw))
        ws = wb["Cronograma"]
        # Header de Lunes NO debería estar merged (1 sola col).
        merges_header = [
            m for m in ws.merged_cells.ranges
            if m.min_row == 1 and m.max_row == 1
            and m.min_col >= 2
        ]
        # Si algún merge existe en fila 1 col B, no debería
        # cubrir más de 1 columna.
        for m in merges_header:
            if m.min_col == 2:
                assert m.max_col == m.min_col

    def test_colores_por_materia_deterministicos(self):
        # Misma materia con dos bloques → mismo color.
        grid = {
            "Lunes": [
                _mk_block("h1", materia_codigo="MAT1", hi=8, hf=10),
            ],
            "Martes": [
                _mk_block("h2", materia_codigo="MAT1", hi=8, hf=10),
            ],
        }
        raw = export_grilla_a_xlsx(
            grid_data=grid, plan_nombre="P", ciclo_label="C",
            filtros={},
        )
        wb = load_workbook(BytesIO(raw))
        ws = wb["Cronograma"]
        # Los bloques ocupan (row 26, col 2) para Lunes y (col 3)
        # para Martes con este layout de slots. Buscamos por
        # contenido en vez de por índice fijo.
        # Recorremos y comparamos fills de celdas con "MAT1".
        celdas_mat1: list = []
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for c in row:
                if c.value and "MAT1" in str(c.value):
                    celdas_mat1.append(c)
        assert len(celdas_mat1) >= 2
        colores = {
            c.fill.fgColor.rgb for c in celdas_mat1 if c.fill.fgColor
        }
        assert len(colores) == 1, (
            f"MAT1 debería tener 1 color; got {colores}"
        )


class TestBuildExportFilename:

    def test_incluye_plan_ciclo_y_fecha(self):
        fn = build_export_filename(
            plan_nombre="Plan v0",
            ciclo_label="2026-1C",
            filtros={},
        )
        assert fn.endswith(".xlsx")
        assert "plan-v0" in fn
        assert "2026-1c" in fn

    def test_incluye_filtros_no_vacios(self):
        fn = build_export_filename(
            plan_nombre="Plan v0",
            ciclo_label="2026-1C",
            filtros={
                "Carrera": "A · Electrónica",
                "Año": "3º",
                "Cuatrimestre": "1C",
                "Alcance": "Sólo comunes",
            },
        )
        assert "electronica" in fn.lower() or "a-" in fn
        assert "3" in fn
        assert "solo-comunes" in fn.lower()

    def test_omite_filtros_en_todas(self):
        fn = build_export_filename(
            plan_nombre="Plan",
            ciclo_label="2026-1C",
            filtros={
                "Carrera": "(sin filtro)",
                "Alcance": "Todas",
            },
        )
        # Ni "sin-filtro" ni "todas" aparecen en el nombre.
        assert "sin-filtro" not in fn.lower()
        assert "todas" not in fn.lower()
