"""Export a Excel del cronograma de la Grilla Horaria del plan.

Genera un ``.xlsx`` con 3 hojas:

1. **Metadata** — plan, ciclo, filtros aplicados, timestamp.
2. **Cronograma** — matriz día × franja al estilo calendario visual.
3. **Detalle** — tabla plana con una fila por horario (materia,
   comisión, día, hora, aula, modalidad, alcance).

Diseñado para consumo por parte de la administración académica al
publicar cronogramas por carrera / año / cuatri o listados de
comunes.
"""

from __future__ import annotations

from datetime import datetime
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment, Border, Font, PatternFill, Side,
)
from openpyxl.utils import get_column_letter

from src.services.schedule_service import ScheduleBlock


DIAS_ORDER = [
    "Lunes", "Martes", "Miércoles", "Jueves", "Viernes", "Sábado",
]


def _fill_header(cell) -> None:
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill(
        "solid", fgColor="1E88E5",
    )
    cell.alignment = Alignment(horizontal="center", vertical="center")


def _fmt_time(t) -> str:
    return t.strftime("%H:%M") if t else ""


def _write_metadata_sheet(
    ws,
    plan_nombre: str,
    ciclo_label: str,
    filtros: dict,
) -> None:
    """Escribe la hoja Metadata con contexto del export."""
    ws.title = "Metadata"
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 60

    ws["A1"] = "Cronograma del plan"
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells("A1:B1")

    ahora = datetime.now().strftime("%Y-%m-%d %H:%M")
    filas = [
        ("Plan", plan_nombre),
        ("Ciclo", ciclo_label),
        ("Exportado", ahora),
        ("", ""),
        ("Filtros aplicados", ""),
    ]
    row = 3
    for k, v in filas:
        ws.cell(row=row, column=1, value=k).font = Font(bold=True)
        ws.cell(row=row, column=2, value=v)
        row += 1

    for label, valor in filtros.items():
        if not valor:
            valor = "(sin filtro)"
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=2, value=valor)
        row += 1


_MATERIA_PALETTE = [
    ("1E88E5", "FFFFFF"),  # azul
    ("43A047", "FFFFFF"),  # verde
    ("F4511E", "FFFFFF"),  # naranja
    ("8E24AA", "FFFFFF"),  # violeta
    ("00897B", "FFFFFF"),  # turquesa
    ("FFB300", "212121"),  # amarillo
    ("3949AB", "FFFFFF"),  # indigo
    ("D81B60", "FFFFFF"),  # magenta
    ("039BE5", "FFFFFF"),  # celeste
    ("7CB342", "FFFFFF"),  # verde lima
    ("6D4C41", "FFFFFF"),  # marrón
    ("546E7A", "FFFFFF"),  # gris azulado
    ("EF6C00", "FFFFFF"),  # naranja oscuro
    ("00ACC1", "FFFFFF"),  # cyan
    ("C0CA33", "212121"),  # lima
    ("5E35B1", "FFFFFF"),  # violeta profundo
    ("2E7D32", "FFFFFF"),  # verde oscuro
    ("AD1457", "FFFFFF"),  # rosa oscuro
    ("00838F", "FFFFFF"),  # verde azulado
    ("F57C00", "FFFFFF"),  # naranja intenso
    ("827717", "FFFFFF"),  # oliva
    ("4527A0", "FFFFFF"),  # púrpura
    ("BF360C", "FFFFFF"),  # rojo tejo
    ("1B5E20", "FFFFFF"),  # verde bosque
]
_EMPTY_FILL_A = "FAFAFA"       # fondo día par
_EMPTY_FILL_B = "F0F0F0"       # fondo día impar (alternado)
_EMPTY_BORDER = "E0E0E0"       # borde fino interior
_HOUR_LINE = "BDBDBD"          # línea suave entre horas completas
_HEADER_FILL = "455A64"        # gris azulado neutro (col y row headers)
_HEADER_TEXT = "FFFFFF"
_BLOCK_BORDER = "37474F"       # borde mediano de cada bloque
_OUTER_BORDER = "212121"       # borde grueso exterior
_DAY_DIVIDER = "37474F"        # divisor grueso entre días


def _hash_indice(*partes: str | int | None) -> int:
    """Índice determinístico dentro de la paleta desde una llave
    compuesta (materia, comisión, etc.).

    Usa MD5 truncado en vez de la suma de ord() clásica para
    minimizar colisiones cuando hay muchas materias distintas —
    la distribución uniforme del hash criptográfico maximiza el uso
    de la paleta.
    """
    import hashlib
    key = "|".join(str(p) for p in partes if p is not None)
    if not key:
        return 0
    h = hashlib.md5(key.encode("utf-8")).digest()
    return int.from_bytes(h[:4], "big") % len(_MATERIA_PALETTE)


def _shade_hex(hex_rgb: str, delta: int) -> str:
    """Ajusta el brillo de un color hex por ``delta`` en cada canal
    RGB. ``delta`` positivo aclara; negativo oscurece. Se satura en
    [0, 255]."""
    r = max(0, min(255, int(hex_rgb[0:2], 16) + delta))
    g = max(0, min(255, int(hex_rgb[2:4], 16) + delta))
    b = max(0, min(255, int(hex_rgb[4:6], 16) + delta))
    return f"{r:02X}{g:02X}{b:02X}"


_SHADE_STEPS = [0, +30, -25, +55, -50, +80, -75]


def _color_para_bloque(
    blk: ScheduleBlock, modo: str,
) -> tuple[str, str]:
    """Asigna un color determinístico al bloque.

    ``modo`` es una de:
      - ``"materia"``: todos los bloques de la misma materia usan
        el mismo color. Simple, útil cuando el plan tiene pocas
        comisiones por materia.
      - ``"materia_comision"``: la materia decide el **color base**
        (mismo tono para todas sus comisiones). La comisión decide
        el **shade** (más claro / mismo / más oscuro) dentro de
        ese tono. Esto evita colisiones entre materias distintas:
        - materias con colores base distintos nunca terminan con
          el mismo color final;
        - materias que colisionan en color base (~ probable con
          muchas materias) usan **offsets de shade personales**:
          cada materia arranca en un shade distinto y avanza por
          comisión, así dos materias con mismo color base + misma
          comisión rara vez terminan en el mismo shade.
    """
    import hashlib
    idx = _hash_indice(blk.materia_codigo)
    bg, fg = _MATERIA_PALETTE[idx]
    if modo == "materia_comision" and blk.comision_numero is not None:
        # Offset propio de la materia dentro del ciclo de shades:
        # segundo hash independiente para desincronizar dos
        # materias que colisionan en color base.
        shift_seed = int.from_bytes(
            hashlib.sha1(
                blk.materia_codigo.encode("utf-8"),
            ).digest()[:4],
            "big",
        ) % len(_SHADE_STEPS)
        delta = _SHADE_STEPS[
            (shift_seed + blk.comision_numero - 1)
            % len(_SHADE_STEPS)
        ]
        if delta != 0:
            bg = _shade_hex(bg, delta)
            # Ajuste de contraste de texto: si el fondo queda muy
            # claro, oscurecer el texto para legibilidad.
            luminancia = (
                int(bg[0:2], 16) * 299
                + int(bg[2:4], 16) * 587
                + int(bg[4:6], 16) * 114
            ) / 1000
            fg = "FFFFFF" if luminancia < 140 else "212121"
    return bg, fg


def _minutos_bloque(blk: ScheduleBlock) -> tuple[int, int]:
    return (
        blk.hora_inicio.hour * 60 + blk.hora_inicio.minute,
        blk.hora_fin.hour * 60 + blk.hora_fin.minute,
    )


def _asignar_lanes(blocks: list[ScheduleBlock]) -> tuple[
    list[tuple[ScheduleBlock, int]], int
]:
    """Asigna a cada bloque un "lane" (sub-columna) sin solapes.

    Es el algoritmo clásico de coloreo greedy para intervalos:
    ordenamos por hora de inicio y asignamos el lane libre más bajo.
    Dos bloques en el mismo lane nunca se solapan.

    Devuelve (lista de (block, lane_idx), cantidad total de lanes).
    """
    if not blocks:
        return [], 1
    ordenados = sorted(blocks, key=lambda b: _minutos_bloque(b))
    lanes_end: list[int] = []  # fin del último bloque en cada lane
    asignaciones: list[tuple[ScheduleBlock, int]] = []
    for blk in ordenados:
        h_s, h_e = _minutos_bloque(blk)
        lane_asignado = None
        for idx, end in enumerate(lanes_end):
            if end <= h_s:
                lane_asignado = idx
                lanes_end[idx] = h_e
                break
        if lane_asignado is None:
            lane_asignado = len(lanes_end)
            lanes_end.append(h_e)
        asignaciones.append((blk, lane_asignado))
    return asignaciones, max(1, len(lanes_end))


def _resumen_bloque(blk: ScheduleBlock) -> str:
    """Texto del contenido de un bloque en la hoja Cronograma."""
    com_tag = f" [C{blk.comision_numero}]" if blk.comision_numero else ""
    lineas = [f"{blk.materia_codigo}{com_tag}"]
    if blk.materia_nombre:
        lineas.append(blk.materia_nombre)
    if blk.virtual:
        lineas.append("💻 Virtual")
    elif blk.aula_label:
        emoji = "🧪" if blk.tipo_clase == "laboratorio" else "📖"
        lineas.append(f"{emoji} {blk.aula_label}")
    # Última línea: rango horario (ej: 09:00 – 11:00) para que el
    # bloque sea legible aún fuera del contexto de la grilla.
    hi = blk.hora_inicio.strftime("%H:%M")
    hf = blk.hora_fin.strftime("%H:%M")
    lineas.append(f"({hi} – {hf})")
    return "\n".join(lineas)


def _write_cronograma_sheet(
    ws,
    grid_data: dict[str, list[ScheduleBlock]],
    color_por: str = "materia",
) -> None:
    """Escribe la hoja Cronograma con una matriz día × franja tipo
    calendario visual, con:

    - **Sub-columnas por día** cuando hay clases paralelas: si un día
      tiene N bloques simultáneos como máximo, se dedica N columnas
      a ese día. Los días sin paralelismo ocupan 1 sola columna.
    - **Merge de celdas** verticalmente para cada bloque a lo largo
      de todas las franjas de 15 min que ocupa (por eso una clase
      de 2h se ve como un rectángulo unido, no como 8 celdas).
    - **Fondo suave** en toda la grilla (celdas vacías) y **color
      por materia** (determinístico por hash del código) en cada
      bloque, con texto blanco/oscuro según luminosidad del fondo.
    """
    ws.title = "Cronograma"

    # Slots de 15 min entre 7:00 y 23:00 (rango operativo típico).
    slots: list[tuple[int, int]] = []
    for h in range(7, 23):
        for m in (0, 15, 30, 45):
            slots.append((h * 60 + m, h * 60 + m + 15))
    n_slots = len(slots)

    def _fmt_slot(a: int, b: int) -> str:
        return (
            f"{a // 60:02d}:{a % 60:02d}–"
            f"{b // 60:02d}:{b % 60:02d}"
        )

    # Paso 1: lanes por día. Cada día tiene N lanes (sub-columnas)
    # según cuántos bloques paralelos tenga como máximo.
    lanes_por_dia: dict[str, list[tuple[ScheduleBlock, int]]] = {}
    ancho_dia: dict[str, int] = {}
    for dia in DIAS_ORDER:
        blocks = grid_data.get(dia, [])
        asign, n_lanes = _asignar_lanes(blocks)
        lanes_por_dia[dia] = asign
        ancho_dia[dia] = n_lanes

    # Paso 2: mapear día → rango de columnas.
    dia_col_start: dict[str, int] = {}
    col = 2  # A es franja
    for dia in DIAS_ORDER:
        dia_col_start[dia] = col
        col += ancho_dia[dia]
    total_cols = col - 1

    # Niveles de borde:
    #   - thin: separadores interiores (celda a celda dentro del cuerpo).
    #   - medium: alrededor de cada bloque de horario.
    #   - thick: separadores estructurales (headers vs cuerpo,
    #     columna A vs cuerpo) y perímetro exterior.
    thin = Side(style="thin", color=_EMPTY_BORDER)
    medium = Side(style="medium", color=_BLOCK_BORDER)
    thick = Side(style="thick", color=_OUTER_BORDER)
    grid_border = Border(
        left=thin, right=thin, top=thin, bottom=thin,
    )
    header_fill = PatternFill("solid", fgColor=_HEADER_FILL)
    header_font = Font(bold=True, color=_HEADER_TEXT, size=11)

    # Paso 3: header (fila 1 = día, con merge sobre sus lanes).
    corner = ws.cell(row=1, column=1, value="Franja")
    corner.font = header_font
    corner.fill = header_fill
    corner.alignment = Alignment(
        horizontal="center", vertical="center",
    )
    for dia in DIAS_ORDER:
        c0 = dia_col_start[dia]
        c1 = c0 + ancho_dia[dia] - 1
        cell = ws.cell(row=1, column=c0, value=dia)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(
            horizontal="center", vertical="center",
        )
        if c1 > c0:
            ws.merge_cells(
                start_row=1, start_column=c0,
                end_row=1, end_column=c1,
            )
            # Rellenar también las celdas ocultas por el merge así
            # el fondo del header cubre todo el ancho.
            for cc in range(c0 + 1, c1 + 1):
                ws.cell(row=1, column=cc).fill = header_fill

    # Paso 4: llenar toda la grilla con fondo alternado por día +
    # label de franja en la columna A. Los bloques se pintan encima
    # en el paso 5. El fondo alternado ayuda a separar visualmente
    # los días cuando la grilla queda densa.
    fill_a = PatternFill("solid", fgColor=_EMPTY_FILL_A)
    fill_b = PatternFill("solid", fgColor=_EMPTY_FILL_B)
    dia_fill: dict[str, PatternFill] = {}
    for idx, dia in enumerate(DIAS_ORDER):
        dia_fill[dia] = fill_a if idx % 2 == 0 else fill_b
    # Mapeo columna → día para pintar rápido.
    col_a_dia: dict[int, str] = {}
    for dia in DIAS_ORDER:
        for offset in range(ancho_dia[dia]):
            col_a_dia[dia_col_start[dia] + offset] = dia

    for i, (a, b) in enumerate(slots, start=2):
        franja_cell = ws.cell(row=i, column=1, value=_fmt_slot(a, b))
        franja_cell.font = Font(
            bold=True, color=_HEADER_TEXT, size=9,
        )
        franja_cell.alignment = Alignment(
            horizontal="right", vertical="center",
        )
        franja_cell.fill = header_fill
        franja_cell.border = grid_border
        for j in range(2, total_cols + 1):
            c = ws.cell(row=i, column=j)
            c.fill = dia_fill[col_a_dia[j]]
            c.border = grid_border

    # Paso 5: pintar cada bloque como un rectángulo unido en su(s)
    # columna(s) y rango de slots. Cada bloque lleva un borde medium
    # para destacarlo del fondo, sin importar el color asignado.
    block_border = Border(
        left=medium, right=medium, top=medium, bottom=medium,
    )
    for dia in DIAS_ORDER:
        for blk, lane in lanes_por_dia[dia]:
            h_s, h_e = _minutos_bloque(blk)
            # Índices de slot [start, end) que cubre el bloque.
            si_start = None
            si_end = None
            for si, (a, b) in enumerate(slots):
                if h_s < b and h_e > a:
                    if si_start is None:
                        si_start = si
                    si_end = si
            if si_start is None:
                continue

            r0 = si_start + 2
            r1 = si_end + 2
            c0 = dia_col_start[dia] + lane
            bg, fg = _color_para_bloque(blk, color_por)
            fill = PatternFill("solid", fgColor=bg)
            texto = _resumen_bloque(blk)
            cell = ws.cell(row=r0, column=c0, value=texto)
            cell.fill = fill
            cell.font = Font(bold=True, color=fg, size=9)
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True,
            )
            cell.border = block_border
            if r1 > r0:
                ws.merge_cells(
                    start_row=r0, start_column=c0,
                    end_row=r1, end_column=c0,
                )
                for r in range(r0 + 1, r1 + 1):
                    _c = ws.cell(row=r, column=c0)
                    _c.fill = fill
                    _c.border = block_border

    # Paso 6: bordes estructurales por encima de los ya asignados.
    #   - Perímetro exterior grueso (fila 1 → última, col A → última).
    #   - Separador grueso entre header (fila 1) y cuerpo (fila 2+).
    #   - Separador grueso entre columna A (franjas) y cuerpo (col B+).
    #
    # openpyxl re-crea el Border al asignar. Para agregar sin perder
    # los bordes finos/medium existentes, componemos con la Border
    # actual de cada celda.
    def _mix(existing: Border, **new_sides) -> Border:
        return Border(
            left=new_sides.get("left", existing.left),
            right=new_sides.get("right", existing.right),
            top=new_sides.get("top", existing.top),
            bottom=new_sides.get("bottom", existing.bottom),
        )

    last_row = n_slots + 1
    last_col = total_cols
    # Perímetro superior (header) e inferior (última fila del cuerpo).
    for j in range(1, last_col + 1):
        top_c = ws.cell(row=1, column=j)
        top_c.border = _mix(top_c.border, top=thick)
        bot_c = ws.cell(row=last_row, column=j)
        bot_c.border = _mix(bot_c.border, bottom=thick)
    # Perímetro izquierdo (col A) y derecho (última col).
    for i in range(1, last_row + 1):
        left_c = ws.cell(row=i, column=1)
        left_c.border = _mix(left_c.border, left=thick)
        right_c = ws.cell(row=i, column=last_col)
        right_c.border = _mix(right_c.border, right=thick)
    # Separador grueso entre header (fila 1) y cuerpo (fila 2).
    for j in range(1, last_col + 1):
        header_c = ws.cell(row=1, column=j)
        header_c.border = _mix(header_c.border, bottom=thick)
        body_top_c = ws.cell(row=2, column=j)
        body_top_c.border = _mix(body_top_c.border, top=thick)
    # Separador grueso entre col A (franjas) y cuerpo (col B).
    for i in range(1, last_row + 1):
        franja_c = ws.cell(row=i, column=1)
        franja_c.border = _mix(franja_c.border, right=thick)
        body_left_c = ws.cell(row=i, column=2)
        body_left_c.border = _mix(body_left_c.border, left=thick)

    # Divisores verticales entre días (borde grueso). Cada día
    # arranca en dia_col_start[dia]; ponemos borde izquierdo grueso
    # ahí (excepto en Lunes que ya recibió el separador de la col A)
    # y borde derecho grueso en la última col del día previo.
    day_divider = Side(style="thick", color=_DAY_DIVIDER)
    for i in range(1, last_row + 1):
        for dia_idx, dia in enumerate(DIAS_ORDER):
            if dia_idx == 0:
                continue  # el borde izquierdo de Lunes ya viene del thick de col A
            c0 = dia_col_start[dia]
            left_c = ws.cell(row=i, column=c0)
            left_c.border = _mix(left_c.border, left=day_divider)
            # También poner right en la última col del día previo,
            # para que si dos días tienen anchos distintos igual se
            # vea el divisor.
            prev_last = dia_col_start[dia] - 1
            if prev_last >= 2:
                prev_c = ws.cell(row=i, column=prev_last)
                prev_c.border = _mix(prev_c.border, right=day_divider)

    # Líneas horarias suaves cada hora completa (07:00, 08:00, …).
    # Marcamos el borde superior de cada slot que arranca en minuto 0.
    hour_line = Side(style="thin", color=_HOUR_LINE)
    for i, (a, _b) in enumerate(slots, start=2):
        if a % 60 == 0:  # slot arranca en hora completa
            for j in range(1, last_col + 1):
                cell = ws.cell(row=i, column=j)
                cell.border = _mix(cell.border, top=hour_line)

    # Anchos de columna: franja angosta, días anchos.
    ws.column_dimensions["A"].width = 12
    for j in range(2, total_cols + 1):
        # Anchura por lane: 22 si el día tiene 1 lane, 18 si tiene
        # más (para que quepan todos sin scroll horizontal).
        ws.column_dimensions[get_column_letter(j)].width = 20
    # Altura mínima de fila para que 3 líneas quepan bien.
    for i in range(2, n_slots + 2):
        ws.row_dimensions[i].height = 18

    ws.freeze_panes = "B2"


def _write_detalle_sheet(
    ws,
    grid_data: dict[str, list[ScheduleBlock]],
) -> None:
    """Tabla plana con una fila por horario, filtrable en Excel."""
    ws.title = "Detalle"
    headers = [
        "Materia (código)",
        "Materia (nombre)",
        "Comisión",
        "Alcance",
        "Día",
        "Hora inicio",
        "Hora fin",
        "Aula",
        "Modalidad",
    ]
    for j, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=j, value=h)
        _fill_header(c)

    # Widths.
    widths = [16, 40, 20, 30, 12, 12, 12, 30, 14]
    for j, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(j)].width = w

    # Recolectar todos los bloques y ordenar por día → hora → materia.
    filas: list[ScheduleBlock] = []
    for dia in DIAS_ORDER:
        for b in grid_data.get(dia, []):
            filas.append(b)

    # (dia, hora_inicio) para ordenar respetando el orden de DIAS_ORDER.
    dia_idx = {d: i for i, d in enumerate(DIAS_ORDER)}
    filas.sort(
        key=lambda b: (
            dia_idx.get(_dia_de_block(b, grid_data), 99),
            b.hora_inicio,
            b.materia_codigo,
        )
    )

    for i, b in enumerate(filas, start=2):
        modalidad = (
            "Virtual" if b.virtual
            else ("Teórica" if b.tipo_clase == "teorica"
                  else ("Laboratorio" if b.tipo_clase == "laboratorio"
                        else "—"))
        )
        aula_txt = (
            "Virtual" if b.virtual
            else (b.aula_label or "Sin aula")
        )
        alcance = b.carreras_label or "—"
        com_txt = (
            b.comision_nombre or
            (f"C{b.comision_numero}" if b.comision_numero else "—")
        )
        ws.cell(row=i, column=1, value=b.materia_codigo)
        ws.cell(row=i, column=2, value=b.materia_nombre)
        ws.cell(row=i, column=3, value=com_txt)
        ws.cell(row=i, column=4, value=alcance)
        ws.cell(row=i, column=5, value=_dia_de_block(b, grid_data))
        ws.cell(row=i, column=6, value=_fmt_time(b.hora_inicio))
        ws.cell(row=i, column=7, value=_fmt_time(b.hora_fin))
        ws.cell(row=i, column=8, value=aula_txt)
        ws.cell(row=i, column=9, value=modalidad)

    # Autofilter en toda la tabla.
    if filas:
        ws.auto_filter.ref = (
            f"A1:{get_column_letter(len(headers))}{len(filas) + 1}"
        )
    ws.freeze_panes = "A2"


def _dia_de_block(
    block: ScheduleBlock,
    grid_data: dict[str, list[ScheduleBlock]],
) -> str:
    """Recupera el día del bloque desde la clave del grid_data.

    `ScheduleBlock` no guarda el día (viene implícito del key del
    dict de la grilla). Al aplanar para la hoja Detalle, necesitamos
    reasociar.
    """
    for dia, blocks in grid_data.items():
        if any(b.entry_id == block.entry_id for b in blocks):
            return dia
    return ""


def export_grilla_a_xlsx(
    grid_data: dict[str, list[ScheduleBlock]],
    plan_nombre: str,
    ciclo_label: str,
    filtros: dict,
    color_por: str = "materia",
) -> bytes:
    """Genera el Excel completo y devuelve los bytes listos para
    ofrecer al usuario via ``st.download_button``.

    Args:
        grid_data: día → lista de ScheduleBlock (lo que la UI ya
            filtró; el export exporta exactamente lo que hay en el
            grid, sin re-filtrar).
        plan_nombre: nombre visible del plan (para la hoja Metadata).
        ciclo_label: label del ciclo (ej. '2026-1C').
        filtros: dict ``{"Nombre filtro": "valor"}`` para la hoja
            Metadata. Ej.: ``{"Carrera": "A · Electrónica", ...}``.
        color_por: modo de coloreo de los bloques en la hoja
            Cronograma. ``"materia"`` (default) usa el mismo color
            para todas las comisiones de una misma materia.
            ``"materia_comision"`` asigna colores distintos a
            comisiones distintas de la misma materia — útil cuando
            hay muchas comisiones y se necesita distinguirlas.

    Returns:
        bytes del archivo .xlsx.
    """
    wb = Workbook()
    default = wb.active
    _write_metadata_sheet(default, plan_nombre, ciclo_label, filtros)
    cronograma_ws = wb.create_sheet("Cronograma")
    _write_cronograma_sheet(cronograma_ws, grid_data, color_por=color_por)
    detalle_ws = wb.create_sheet("Detalle")
    _write_detalle_sheet(detalle_ws, grid_data)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_export_filename(
    plan_nombre: str,
    ciclo_label: str,
    filtros: dict,
) -> str:
    """Nombre de archivo sugerido para el download.

    Formato: ``plan_ciclo_ubicacion_alcance_YYYYMMDD.xlsx``.
    Los filtros que estén en 'Todas'/vacío se omiten.
    """
    fecha = datetime.now().strftime("%Y%m%d")
    partes = [_slug(plan_nombre), _slug(ciclo_label)]
    car = filtros.get("Carrera")
    anio = filtros.get("Año")
    cuatri = filtros.get("Cuatrimestre")
    if car and car != "(sin filtro)":
        partes.append(_slug(str(car)))
    if anio and anio != "(sin filtro)":
        partes.append(_slug(str(anio)))
    if cuatri and cuatri != "(sin filtro)":
        partes.append(_slug(str(cuatri)))
    alcance = filtros.get("Alcance")
    if alcance and alcance not in ("Todas", "(sin filtro)"):
        partes.append(_slug(str(alcance)))
    partes.append(fecha)
    base = "_".join(p for p in partes if p)
    return f"{base}.xlsx"


def _slug(s: str) -> str:
    import re
    import unicodedata

    # Normalizar acentos y símbolos unicode a ASCII equivalente para
    # que el nombre del archivo sea seguro en cualquier filesystem.
    s = unicodedata.normalize("NFKD", s)
    s = "".join(c for c in s if not unicodedata.combining(c))
    s = s.strip().replace(" ", "-").replace("·", "").replace(",", "")
    s = re.sub(r"[^\w\-]", "", s)
    return re.sub(r"-+", "-", s).strip("-").lower()
